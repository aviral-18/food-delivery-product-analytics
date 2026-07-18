"""Report exports: CSV, Excel (xlsx), and PDF."""
from __future__ import annotations

import csv
import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.analytics import catalog, customers, marketing
from app.analytics.filters import Filters
from app.api.deps import get_current_user, get_filters
from app.core.database import get_db
from app.models import User
from app.services.audit import log_action

router = APIRouter()


def _resolve_report(db: Session, f: Filters, report: str) -> tuple[str, list[dict]]:
    """Return (title, rows) for a report key."""
    if report == "city-performance":
        return "City Performance", catalog.city_performance(db, f)["cities"]
    if report == "restaurant-performance":
        return "Restaurant Performance", catalog.restaurant_performance(db, f, limit=500)["restaurants"]
    if report == "cuisine-performance":
        return "Cuisine Performance", catalog.cuisine_performance(db, f)["cuisines"]
    if report == "coupon-effectiveness":
        return "Coupon Effectiveness", marketing.coupon_effectiveness(db, f)["coupons"]
    if report == "rfm-segments":
        return "RFM Segments", customers.rfm_segmentation(db, f)["segments"]
    if report == "clv-deciles":
        return "CLV Deciles", customers.clv_analysis(db, f)["deciles"]
    raise HTTPException(status.HTTP_404_NOT_FOUND, f"Unknown report: {report}")


_REPORTS = ["city-performance", "restaurant-performance", "cuisine-performance",
            "coupon-effectiveness", "rfm-segments", "clv-deciles"]


@router.get("/reports")
def available_reports(_: User = Depends(get_current_user)):
    return {"reports": _REPORTS, "formats": ["csv", "xlsx", "pdf"]}


def _audit(db, user, request, report, fmt):
    log_action(db, user_id=user.id, action="export", entity="report", entity_id=report,
               ip_address=request.client.host if request.client else None, detail=f"format={fmt}")


@router.get("/{report}.csv")
def export_csv(report: str, request: Request, f: Filters = Depends(get_filters),
               db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    title, rows = _resolve_report(db, f, report)
    buf = io.StringIO()
    if rows:
        writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    _audit(db, user, request, report, "csv")
    return StreamingResponse(
        iter([buf.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={report}.csv"},
    )


@router.get("/{report}.xlsx")
def export_xlsx(report: str, request: Request, f: Filters = Depends(get_filters),
                db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    title, rows = _resolve_report(db, f, report)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
        for row in rows:
            ws.append([row.get(h) for h in headers])
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 3, 40)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    _audit(db, user, request, report, "xlsx")
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={report}.xlsx"},
    )


@router.get("/{report}.pdf")
def export_pdf(report: str, request: Request, f: Filters = Depends(get_filters),
               db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet

    title, rows = _resolve_report(db, f, report)
    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4), topMargin=1.2 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Eternal Product Analytics — {title}", styles["Title"]),
        Paragraph(f"Generated {datetime.now():%Y-%m-%d %H:%M} · period {f.effective_start} to {f.effective_end}", styles["Normal"]),
        Spacer(1, 0.4 * cm),
    ]
    if rows:
        headers = list(rows[0].keys())[:8]  # keep it printable
        data = [headers] + [[_short(r.get(h)) for h in headers] for r in rows[:40]]
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ]))
        story.append(table)
    else:
        story.append(Paragraph("No data for the selected filters.", styles["Normal"]))
    doc.build(story)
    out.seek(0)
    _audit(db, user, request, report, "pdf")
    return StreamingResponse(
        out, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={report}.pdf"},
    )


def _short(v) -> str:
    if isinstance(v, float):
        return f"{v:,.2f}"
    s = str(v) if v is not None else ""
    return s[:22]
